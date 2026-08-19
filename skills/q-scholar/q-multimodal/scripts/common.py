"""
Shared utilities for multimodal analysis scripts.

Provides common functions used across pillow and opensmile pipelines:
read_input, save_excel, derive_subject, merge_checkpoints.
"""

import os
import re
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def read_input(path, str_cols=None):
    """Read tabular input file, auto-detecting format by extension.

    str_cols are read as strings, never through numeric inference — a 19-digit
    id inferred as a number is silently rounded at the Excel float layer."""
    conv = {c: str for c in (str_cols or [])}
    ext = Path(path).suffix.lower()
    if ext in (".xlsx", ".xls"):
        df = pd.read_excel(path, converters=conv)
    elif ext == ".csv":
        df = pd.read_csv(path, converters=conv)
    elif ext == ".json":
        df = pd.read_json(path)
    elif ext == ".parquet":
        df = pd.read_parquet(path)
    elif ext == ".tsv":
        df = pd.read_csv(path, sep="\t", converters=conv)
    else:
        raise ValueError(f"Unsupported input format: {ext}. Use xlsx, csv, json, parquet, or tsv.")
    for c in (str_cols or []):
        if c in df.columns:
            df[c] = df[c].map(lambda v: v if pd.isna(v) else str(v))
    return df


def save_excel(df, path):
    """Save DataFrame to Excel with bold headers, auto-fit widths, frozen panes."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="features")
        ws = writer.sheets["features"]

        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="bottom")

        for col_idx, col_name in enumerate(df.columns, 1):
            letter = get_column_letter(col_idx)
            header_cell = ws[f"{letter}1"]
            header_cell.font = header_font
            header_cell.alignment = header_alignment

            sample = df[col_name].head(100).dropna().astype(str)
            max_val_len = int(sample.str.len().max()) if len(sample) else 0
            width = min(max(len(col_name), min(max_val_len, 50)) + 2, 55)
            ws.column_dimensions[letter].width = width

        ws.freeze_panes = "A2"


def source_columns(id_cols, file_col, extra=None):
    """Output/merge key columns: id_cols + file_col (always retained, so each
    row identifies its exact asset) + extra (e.g. frame_number). Ordered,
    deduplicated."""
    keep = list(id_cols) if id_cols else []
    if file_col not in keep:
        keep.append(file_col)
    for c in (extra or []):
        if c not in keep:
            keep.append(c)
    seen = set()
    return [c for c in keep if not (c in seen or seen.add(c))]


def derive_subject(row, file_col, group_col):
    """Derive subject name from group column or file path parent directory."""
    if group_col:
        raw = row.get(group_col, None)
        if raw is None or (isinstance(raw, float) and raw != raw):
            val = "unknown"
        else:
            val = str(raw)
        val = val.replace("\\", "/").rstrip("/")
        val = val.split("/")[-1] if "/" in val else val
    else:
        fp = str(row.get(file_col, ""))
        parts = fp.replace("\\", "/").rstrip("/").split("/")
        val = parts[-2] if len(parts) >= 2 else "default"
    return re.sub(r'[<>:"/\\|?*]', '_', val).strip('. ') or "default"


def merge_checkpoints(checkpoint_dir, output_path, file_col="file_path",
                      exclude_prefix="_", dedup_cols=None):
    """Merge all checkpoint xlsx files in a directory into one file.

    Args:
        checkpoint_dir: Directory containing per-subject checkpoint xlsx files.
        output_path: Path for the merged output xlsx file.
        file_col: Column name used as default dedup key (default: "file_path").
        exclude_prefix: Skip files whose name starts with this prefix.
        dedup_cols: Columns for deduplication. Defaults to [file_col].
            Also read as strings (see read_input).
    Returns:
        Tuple of (merged_df, stats_dict) with keys: files, rows, deduped.
    """
    empty_stats = {"files": 0, "rows": 0, "deduped": 0}
    ckpt_path = Path(checkpoint_dir)
    if not ckpt_path.is_dir():
        print(f"  No checkpoint directory: {checkpoint_dir}", flush=True)
        return pd.DataFrame(), empty_stats

    files = sorted(f for f in ckpt_path.glob("*.xlsx") if not f.name.startswith(exclude_prefix))
    if not files:
        print(f"  No checkpoints found in {checkpoint_dir}", flush=True)
        return pd.DataFrame(), empty_stats

    # Fail closed: a skipped checkpoint, partial key, or schema union would
    # silently drop, collapse, or null-pad assets.
    key_cols = dedup_cols if dedup_cols is not None else [file_col]
    conv = {c: str for c in key_cols if c != "frame_number"}
    dfs, errors, schema = [], [], None
    for f in files:
        try:
            df = pd.read_excel(f, converters=conv)
        except Exception as e:
            errors.append(f"{f.name}: unreadable ({e})")
            continue
        dupes = df.columns[df.columns.duplicated()].tolist()
        if dupes:
            errors.append(f"{f.name}: duplicate column names {dupes}")
            continue
        missing = [c for c in key_cols if c not in df.columns]
        if missing:
            errors.append(f"{f.name}: missing key column(s) {missing}")
            continue
        cols = list(df.columns)
        if schema is None:
            schema = (f.name, cols)
        elif cols != schema[1]:
            lacks = [c for c in schema[1] if c not in cols]
            extra = [c for c in cols if c not in schema[1]]
            detail = (f"missing {lacks}, extra {extra}" if lacks or extra
                      else "same columns in a different order")
            errors.append(f"{f.name}: schema differs from {schema[0]} ({detail})")
            continue
        dfs.append(df)
    if errors:
        raise RuntimeError(
            "checkpoint merge aborted; fix these checkpoints and re-run:\n  "
            + "\n  ".join(errors))

    if not dfs:
        return pd.DataFrame(), empty_stats

    merged = pd.concat(dfs, ignore_index=True)
    before = len(merged)
    merged = merged.drop_duplicates(subset=key_cols, keep="last")
    deduped = before - len(merged)

    out = Path(output_path)
    tmp = out.with_name(out.name + ".tmp")
    save_excel(merged, tmp)
    os.replace(tmp, out)
    return merged, {"files": len(dfs), "rows": len(merged), "deduped": deduped}
