"""
run_eda.py — Universal Exploratory Data Analysis (EDA) Runner
q-exploratory-analysis sub-skill of q-scholar

Six-phase pipeline applying measurement-level-appropriate analysis
based on Stevens' levels (Nominal, Ordinal, Discrete, Continuous)
plus Temporal, Text, and ID/key.

Usage:
    python scripts/run_eda.py data.xlsx --output TABLE/
    python scripts/run_eda.py data.xlsx --col_types rating=ordinal description=text --group platform tier --output TABLE/
    python scripts/run_eda.py data.xlsx --no_excel --output TABLE/
"""

import argparse
import glob
import os
import re
import sys
from collections import Counter
from itertools import combinations

import numpy as np
import pandas as pd
from scipy import stats


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

STOPWORDS = {
    "a", "an", "the", "and", "or", "but", "in", "on", "at", "to", "for",
    "of", "with", "by", "from", "is", "are", "was", "were", "be", "been",
    "being", "have", "has", "had", "do", "does", "did", "will", "would",
    "shall", "should", "may", "might", "must", "can", "could", "that",
    "this", "these", "those", "it", "its", "i", "you", "he", "she", "we",
    "they", "their", "them", "his", "her", "our", "your", "my", "not",
    "no", "so", "if", "as", "up", "out", "about", "into", "than", "then",
    "there", "what", "which", "who", "when", "where", "how", "all", "each",
}

LOW_CARD_MAX = 20  # integers with <= this many unique values are flagged as ambiguous


# ---------------------------------------------------------------------------
# Phase 0: Column Classification
# ---------------------------------------------------------------------------

def classify_columns(df: pd.DataFrame,
                     col_types_override: dict = None) -> dict:
    """
    Auto-detect measurement level for each column.
    Returns dict {col_name: type_string}.

    Type strings: 'id', 'binary', 'nominal', 'ordinal', 'discrete',
                  'continuous', 'temporal', 'text', 'unknown'
    """
    col_types_override = col_types_override or {}
    col_types = {}
    n = len(df)

    for col in df.columns:
        series = df[col]

        # User-confirmed type (highest priority)
        if col in col_types_override:
            col_types[col] = col_types_override[col]
            continue

        nunique = series.nunique(dropna=True)

        # Datetime
        if pd.api.types.is_datetime64_any_dtype(series):
            col_types[col] = "temporal"
            continue

        # Try parsing as datetime if object type
        if pd.api.types.is_object_dtype(series):
            sample = series.dropna().head(20)
            parsed = pd.to_datetime(sample, errors="coerce")
            if parsed.notna().mean() >= 0.8:
                col_types[col] = "temporal"
                continue

        # ID/key: nearly all unique values
        if nunique > 0.95 * n and n > 10:
            if pd.api.types.is_numeric_dtype(series):
                # Numeric columns with near-unique values are metrics, not IDs
                col_types[col] = "continuous"
            else:
                col_types[col] = "id"
            continue

        # Binary: exactly 2 unique values
        if nunique == 2:
            col_types[col] = "binary"
            continue

        # Text: object with long average string length
        if pd.api.types.is_object_dtype(series):
            avg_len = series.dropna().astype(str).str.len().mean()
            if avg_len > 50:
                col_types[col] = "text"
            else:
                col_types[col] = "nominal"
            continue

        # Numeric types
        if pd.api.types.is_integer_dtype(series) or (
            pd.api.types.is_float_dtype(series) and (series.dropna() % 1 == 0).all()
        ):
            if nunique <= LOW_CARD_MAX:
                # Ambiguous: could be ordinal (Likert) or discrete (small count)
                # Defaults to 'discrete' unless --interactive is set
                col_types[col] = "ambiguous_integer"
            else:
                col_types[col] = "discrete"
            continue

        if pd.api.types.is_float_dtype(series):
            col_types[col] = "continuous"
            continue

        col_types[col] = "unknown"

    return col_types


def resolve_ambiguous_columns(col_types: dict, df: pd.DataFrame,
                               no_interactive: bool = False) -> dict:
    """
    Classify ambiguous integer columns as 'ordinal' or 'discrete'.

    If no_interactive=True or stdin is not a TTY, auto-classifies as 'discrete'
    (conservative default) with a printed notice. Use --col_types col=ordinal to
    pre-specify known ordinal columns and avoid this auto-classification.
    Interactive prompt fires only when no_interactive=False and stdin is a TTY.
    """
    ambiguous = [c for c, t in col_types.items() if t == "ambiguous_integer"]
    if not ambiguous:
        return col_types

    if no_interactive or not sys.stdin.isatty():
        print("\n--- Ambiguous Integer Columns: Auto-Classification ---")
        for col in ambiguous:
            col_types[col] = "discrete"
            vals = sorted(df[col].dropna().unique().tolist())
            print(
                f"  Column '{col}' ({len(vals)} unique values) -> classified as 'discrete'. "
                f"Use --col_types {col}=ordinal to override."
            )
        return col_types

    print("\n--- Ambiguous Integer Columns Detected ---")
    print("The following integer columns have low cardinality and could be")
    print("Ordinal (e.g., Likert scale) or Discrete (e.g., count).\n")

    for col in ambiguous:
        vals = sorted(df[col].dropna().unique().tolist())
        print(f"  Column: '{col}'")
        print(f"  Unique values ({len(vals)}): {vals[:10]}{'...' if len(vals) > 10 else ''}")
        while True:
            choice = input("  Classify as [o]rdinal or [d]iscrete? ").strip().lower()
            if choice in ("o", "ordinal"):
                col_types[col] = "ordinal"
                break
            elif choice in ("d", "discrete"):
                col_types[col] = "discrete"
                break
            else:
                print("  Please enter 'o' or 'd'.")
        print()

    return col_types


def print_schema_summary(df: pd.DataFrame, col_types: dict) -> None:
    """Print schema summary table to stdout."""
    print("\n=== Phase 0: Column Classification ===")
    print(f"{'Column':<30} {'Type':<14} {'Nunique':>8} {'Missing%':>9}")
    print("-" * 65)
    for col in df.columns:
        t = col_types.get(col, "unknown")
        nunique = df[col].nunique(dropna=True)
        missing_pct = df[col].isna().mean() * 100
        print(f"{col:<30} {t:<14} {nunique:>8} {missing_pct:>8.1f}%")
    print()


# ---------------------------------------------------------------------------
# Phase 1: Dataset Profile
# ---------------------------------------------------------------------------

def dataset_profile(df: pd.DataFrame, col_types: dict) -> pd.DataFrame:
    n_rows, n_cols = df.shape
    rows = []
    for col in df.columns:
        rows.append({
            "column": col,
            "detected_type": col_types.get(col, "unknown"),
            "dtype": str(df[col].dtype),
            "nunique": df[col].nunique(dropna=True),
            "missing_count": df[col].isna().sum(),
            "missing_pct": round(df[col].isna().mean() * 100, 2),
            "memory_kb": round(df[col].memory_usage(deep=True) / 1024, 2),
        })
    profile = pd.DataFrame(rows)
    profile.insert(0, "n_rows", n_rows)
    profile.insert(1, "n_cols", n_cols)
    return profile


# ---------------------------------------------------------------------------
# Phase 2: Data Quality
# ---------------------------------------------------------------------------

def check_data_quality(df: pd.DataFrame, col_types: dict) -> pd.DataFrame:
    rows = []
    dup_count = df.duplicated().sum()

    for col in df.columns:
        t = col_types.get(col, "unknown")
        series = df[col]
        missing_count = series.isna().sum()
        missing_pct = round(series.isna().mean() * 100, 2)
        is_constant = series.nunique(dropna=True) <= 1

        mild_outliers = 0
        extreme_outliers = 0
        if t in ("continuous", "discrete"):
            num = series.dropna()
            if len(num) > 3:
                q1 = num.quantile(0.25)
                q3 = num.quantile(0.75)
                iqr = q3 - q1
                mild_outliers = int(((num < q1 - 1.5 * iqr) | (num > q3 + 1.5 * iqr)).sum())
                extreme_outliers = int(((num < q1 - 3.0 * iqr) | (num > q3 + 3.0 * iqr)).sum())

        rows.append({
            "column": col,
            "type": t,
            "missing_count": missing_count,
            "missing_pct": missing_pct,
            "duplicate_rows": dup_count,
            "is_constant": is_constant,
            "mild_outliers_iqr1.5": mild_outliers,
            "extreme_outliers_iqr3.0": extreme_outliers,
        })

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Phase 3 helpers: Univariate Analysis
# ---------------------------------------------------------------------------

def quantitative_summary(series: pd.Series, name: str, level: str) -> dict:
    """
    Unified quantitative metrics for Ordinal, Discrete, and Continuous columns.
    level: 'ordinal' | 'discrete' | 'continuous'
    """
    s = series.dropna()
    n = len(s)
    if n == 0:
        return {"variable": name, "level": level, "N_valid": 0}
    if n == 1:
        mean_label = "M_quasi_interval" if level == "ordinal" else "M"
        return {"variable": name, "level": level, "N_valid": 1,
                mean_label: s.iloc[0], "Mdn": s.iloc[0],
                "note": "n=1; dispersion metrics unavailable"}

    mean = s.mean()
    median = s.median()
    try:
        mode_val = s.mode().iloc[0]
    except IndexError:
        mode_val = np.nan
    sd = s.std(ddof=1)
    variance = s.var(ddof=1)
    minimum = s.min()
    maximum = s.max()
    range_val = maximum - minimum
    q1 = s.quantile(0.25)
    q3 = s.quantile(0.75)
    iqr = q3 - q1
    cv = (sd / mean * 100) if mean != 0 else np.nan
    skewness = s.skew()
    kurt = s.kurt()
    se = sd / np.sqrt(n)
    ci_lower = mean - 1.96 * se
    ci_upper = mean + 1.96 * se
    missing = series.isna().sum()
    missing_pct = round(series.isna().mean() * 100, 2)

    mild_outliers = int(((s < q1 - 1.5 * iqr) | (s > q3 + 1.5 * iqr)).sum())
    extreme_outliers = int(((s < q1 - 3.0 * iqr) | (s > q3 + 3.0 * iqr)).sum())

    mean_label = "M_quasi_interval" if level == "ordinal" else "M"

    result = {
        "variable": name,
        "level": level,
        "N_valid": n,
        "missing_count": missing,
        "missing_pct": missing_pct,
        mean_label: round(mean, 4),
        "Mdn": round(median, 4),
        "Mode": mode_val,
        "SD": round(sd, 4),
        "Variance": round(variance, 4),
        "Range": round(range_val, 4),
        "IQR": round(iqr, 4),
        "CV_pct": round(cv, 2) if not np.isnan(cv) else np.nan,
        "Min": minimum,
        "Q1_25th": round(q1, 4),
        "Q3_75th": round(q3, 4),
        "Max": maximum,
        "Skewness": round(skewness, 4),
        "Kurtosis": round(kurt, 4),
        "SE": round(se, 4),
        "CI95_lower": round(ci_lower, 4),
        "CI95_upper": round(ci_upper, 4),
        "mild_outliers_IQR1.5": mild_outliers,
        "extreme_outliers_IQR3.0": extreme_outliers,
    }

    if level in ("continuous", "discrete"):
        result["P10_10th"] = round(s.quantile(0.10), 4)
        result["P90_90th"] = round(s.quantile(0.90), 4)

    return result


def nominal_frequencies(df: pd.DataFrame, col_types: dict, top_n: int = 10) -> pd.DataFrame:
    nominal_cols = [c for c, t in col_types.items() if t == "nominal"]
    rows = []
    for col in nominal_cols:
        vc = df[col].value_counts(dropna=True).head(top_n)
        total = df[col].notna().sum()
        for val, count in vc.items():
            rows.append({
                "column": col,
                "value": val,
                "count": count,
                "pct": round(count / total * 100, 2) if total > 0 else 0,
            })
    return pd.DataFrame(rows)


def binary_summary(df: pd.DataFrame, col_types: dict) -> pd.DataFrame:
    binary_cols = [c for c, t in col_types.items() if t == "binary"]
    rows = []
    for col in binary_cols:
        s = df[col].dropna()
        n = len(s)
        vals = s.value_counts()
        for val, count in vals.items():
            p = count / n
            se = np.sqrt(p * (1 - p) / n)
            rows.append({
                "column": col,
                "value": val,
                "count": count,
                "proportion": round(p, 4),
                "CI95_lower": round(max(0.0, p - 1.96 * se), 4),
                "CI95_upper": round(min(1.0, p + 1.96 * se), 4),
            })
    return pd.DataFrame(rows)


def ordinal_frequency(df: pd.DataFrame, col_types: dict) -> pd.DataFrame:
    ordinal_cols = [c for c, t in col_types.items() if t == "ordinal"]
    rows = []
    for col in ordinal_cols:
        s = df[col].dropna().sort_values()
        total = len(s)
        vc = s.value_counts().sort_index()
        cumulative = 0
        for val, count in vc.items():
            cumulative += count
            rows.append({
                "column": col,
                "value": val,
                "count": count,
                "pct": round(count / total * 100, 2) if total > 0 else 0,
                "cumulative_pct": round(cumulative / total * 100, 2) if total > 0 else 0,
            })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Phase 4: Bivariate Analysis
# ---------------------------------------------------------------------------

def pearson_correlation_matrix(df: pd.DataFrame, col_types: dict) -> pd.DataFrame:
    # Include both continuous AND discrete for Pearson (both are ratio-scale)
    ratio_cols = [c for c, t in col_types.items() if t in ("continuous", "discrete")]
    if len(ratio_cols) < 2:
        return pd.DataFrame()

    sub = df[ratio_cols].dropna()
    rows = []
    for c1, c2 in combinations(ratio_cols, 2):
        s1, s2 = sub[c1], sub[c2]
        n = len(s1)
        if n < 3:
            continue
        r, p = stats.pearsonr(s1, s2)
        rows.append({
            "col1": c1, "col2": c2,
            "r": round(r, 4), "p_value": round(p, 4), "N": n,
        })
    return pd.DataFrame(rows)


def spearman_correlation_matrix(df: pd.DataFrame, col_types: dict) -> pd.DataFrame:
    ord_cols = [c for c, t in col_types.items() if t == "ordinal"]
    if len(ord_cols) < 2:
        return pd.DataFrame()

    sub = df[ord_cols].dropna()
    rows = []
    for c1, c2 in combinations(ord_cols, 2):
        s1, s2 = sub[c1], sub[c2]
        n = len(s1)
        if n < 3:
            continue
        rho, p = stats.spearmanr(s1, s2)
        rows.append({
            "col1": c1, "col2": c2,
            "rho": round(rho, 4), "p_value": round(p, 4), "N": n,
        })
    return pd.DataFrame(rows)


def grouped_stats_by_nominal(df: pd.DataFrame, measure_col: str,
                              group_col: str, measure_type: str) -> pd.DataFrame:
    """
    Grouped descriptives for a measure column stratified by a nominal group column.
    measure_type: 'continuous' | 'discrete' | 'ordinal'
    """
    rows = []
    for group_val in df[group_col].dropna().unique():
        sub = df[df[group_col] == group_val][measure_col].dropna()
        if len(sub) == 0:
            continue
        row = {"group_col": group_col, "group_value": group_val,
               "measure_col": measure_col, "N": len(sub)}
        if measure_type in ("continuous", "discrete"):
            row["M"] = round(sub.mean(), 4)
            row["SD"] = round(sub.std(ddof=1), 4)
            row["Mdn"] = round(sub.median(), 4)
            row["IQR"] = round(sub.quantile(0.75) - sub.quantile(0.25), 4)
        elif measure_type == "ordinal":
            row["Mdn"] = round(sub.median(), 4)
            row["IQR"] = round(sub.quantile(0.75) - sub.quantile(0.25), 4)
        rows.append(row)
    return pd.DataFrame(rows)


def cross_tabulate(df: pd.DataFrame, col1: str, col2: str) -> pd.DataFrame:
    ct = pd.crosstab(df[col1], df[col2], margins=True, margins_name="Total")
    return ct.reset_index()


# ---------------------------------------------------------------------------
# Phase 5: Specialized Analysis
# ---------------------------------------------------------------------------

def text_analysis(series: pd.Series, col_name: str, top_n: int = 20) -> pd.DataFrame:
    texts = series.dropna().astype(str).tolist()
    if not texts:
        return pd.DataFrame()

    word_counts = [len(t.split()) for t in texts]
    all_words = []
    all_bigrams = []

    for text in texts:
        tokens = [w.lower() for w in re.findall(r"\b[a-zA-Z]+\b", text)
                  if w.lower() not in STOPWORDS and len(w) > 1]
        all_words.extend(tokens)
        all_bigrams.extend([f"{tokens[i]} {tokens[i+1]}" for i in range(len(tokens) - 1)])

    word_freq = Counter(all_words)
    bigram_freq = Counter(all_bigrams)
    vocab_size = len(word_freq)

    rows = []
    for w, c in word_freq.most_common(top_n):
        rows.append({"column": col_name, "type": "unigram", "item": w, "count": c})
    for bg, c in bigram_freq.most_common(top_n):
        rows.append({"column": col_name, "type": "bigram", "item": bg, "count": c})

    summary_rows = [
        {"column": col_name, "type": "summary", "item": "avg_word_count",
         "count": round(np.mean(word_counts), 2)},
        {"column": col_name, "type": "summary", "item": "vocab_size",
         "count": vocab_size},
        {"column": col_name, "type": "summary", "item": "total_docs",
         "count": len(texts)},
    ]
    return pd.DataFrame(summary_rows + rows)


def temporal_trends(df: pd.DataFrame, col_types: dict,
                    measure_cols: list = None) -> pd.DataFrame:
    temporal_cols = [c for c, t in col_types.items() if t == "temporal"]
    if not temporal_cols:
        return pd.DataFrame()

    temp_col = temporal_cols[0]
    dates = pd.to_datetime(df[temp_col], errors="coerce")
    df = df.copy()
    df["_period"] = dates.dt.to_period("M").astype(str)

    if measure_cols is None:
        measure_cols = [c for c, t in col_types.items()
                        if t in ("continuous", "discrete")]

    rows = []
    for period, grp in df.groupby("_period"):
        row = {"period": period, "count": len(grp)}
        for col in measure_cols:
            num = grp[col].dropna()
            if len(num) > 0:
                row[f"{col}_mean"] = round(num.mean(), 4)
                row[f"{col}_median"] = round(num.median(), 4)
        rows.append(row)
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Phase 6: Excel Report
# ---------------------------------------------------------------------------

def generate_excel_report(output_dir: str) -> str:
    """
    Generate EXPLORATORY_REPORT.xlsx with APA-7 formatting (B&W, no color fills).
    Returns path to the generated file.

    Structure:
    - Sheet 0 "Summary": dataset dimensions, type counts, quality highlights
    - Sheets 1-N: one per generated CSV (only files that exist are included)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    thin = Side(border_style="thin", color="000000")

    def _set_col_widths(ws, df: pd.DataFrame) -> None:
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            header_len = len(str(col_name))
            max_data_len = (int(df[col_name].astype(str).str.len().max())
                            if len(df) > 0 else 0)
            width = min(max(header_len, max_data_len) + 2, 40)
            ws.column_dimensions[col_letter].width = width

    def _write_apa_table(ws, table_num: int, title: str,
                          df: pd.DataFrame, note: str = "") -> None:
        ws.sheet_view.showGridLines = False
        row_cursor = 1

        # "Table N" row -- bold
        cell = ws.cell(row=row_cursor, column=1, value=f"Table {table_num}")
        cell.font = Font(name="Calibri", size=11, bold=True)
        row_cursor += 1

        # Title row -- sentence-case italic
        cell = ws.cell(row=row_cursor, column=1, value=title)
        cell.font = Font(name="Calibri", size=11, italic=True)
        row_cursor += 2  # blank separator row

        # Column header row -- bold, single bottom border
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=row_cursor, column=col_idx, value=str(col_name))
            cell.font = Font(name="Calibri", size=11, bold=True)
            cell.border = Border(bottom=thin)
            is_numeric = pd.api.types.is_numeric_dtype(df[col_name])
            cell.alignment = Alignment(horizontal="right" if is_numeric else "left")
        row_cursor += 1

        # Data rows -- plain; thin top on first row, thin bottom on last row only
        for ridx, (_, data_row) in enumerate(df.iterrows()):
            is_first = ridx == 0
            is_last = ridx == len(df) - 1
            for col_idx, col_name in enumerate(df.columns, start=1):
                val = data_row[col_name]
                if pd.isna(val):
                    val = ""
                cell = ws.cell(row=row_cursor, column=col_idx, value=val)
                cell.font = Font(name="Calibri", size=11)
                if is_first and is_last:
                    cell.border = Border(top=thin, bottom=thin)
                elif is_first:
                    cell.border = Border(top=thin)
                elif is_last:
                    cell.border = Border(bottom=thin)
                is_numeric_val = isinstance(val, (int, float)) and val != ""
                cell.alignment = Alignment(
                    horizontal="right" if is_numeric_val else "left"
                )
            row_cursor += 1

        # Note row -- "Note." italic
        if note:
            row_cursor += 1
            note_cell = ws.cell(row=row_cursor, column=1, value=f"Note. {note}")
            note_cell.font = Font(name="Calibri", size=11, italic=True)

        _set_col_widths(ws, df)

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ---- Sheet 0: Summary ----
    ws_summary = wb.create_sheet(title="Summary")
    summary_rows: list = []

    profile_path = os.path.join(output_dir, "01_dataset_profile.csv")
    dq_path = os.path.join(output_dir, "02_data_quality.csv")

    if os.path.isfile(profile_path):
        df_profile = pd.read_csv(profile_path)
        if not df_profile.empty:
            n_rows_val = int(df_profile["n_rows"].iloc[0])
            n_cols_val = int(df_profile["n_cols"].iloc[0])
            summary_rows.append({"Metric": "Total rows", "Value": f"{n_rows_val:,}"})
            summary_rows.append({"Metric": "Total columns", "Value": str(n_cols_val)})
            for t, cnt in df_profile["detected_type"].value_counts().items():
                summary_rows.append({"Metric": f"Columns ({t})", "Value": str(cnt)})

    if os.path.isfile(dq_path):
        df_dq = pd.read_csv(dq_path)
        if not df_dq.empty:
            dup = (int(df_dq["duplicate_rows"].iloc[0])
                   if "duplicate_rows" in df_dq.columns else 0)
            summary_rows.append({"Metric": "Duplicate rows", "Value": str(dup)})
            for _, r in df_dq[df_dq["missing_pct"] > 20].iterrows():
                summary_rows.append({
                    "Metric": f"High missing: {r['column']}",
                    "Value": f"{r['missing_pct']:.1f}%",
                })

    summary_df = (pd.DataFrame(summary_rows) if summary_rows
                  else pd.DataFrame([{"Metric": "No profile data", "Value": ""}]))
    _write_apa_table(
        ws_summary, 1,
        "Dataset summary and data quality highlights",
        summary_df,
        note=(
            "Column types: nominal, ordinal, discrete, continuous, binary, "
            "temporal, text, id. High missing = >20% missing values."
        ),
    )

    # ---- Sheets 1-N: one per static CSV ----
    static_sheets = [
        ("01_dataset_profile.csv", "01 Profile",
         "Dataset column profile"),
        ("02_data_quality.csv", "02 Quality",
         "Data quality assessment per column"),
        ("03_nominal_frequencies.csv", "03 Nominal",
         "Nominal variable frequency distributions (top-N values)"),
        ("04_binary_summary.csv", "04 Binary",
         "Binary variable proportions and 95% confidence intervals"),
        ("05_ordinal_distribution.csv", "05 Ordinal Dist",
         "Ordinal variable ordered frequency distributions"),
        ("06_ordinal_descriptives.csv", "06 Ordinal Desc",
         "Ordinal variable descriptive statistics (quasi-interval mean)"),
        ("07_discrete_descriptives.csv", "07 Discrete",
         "Discrete variable descriptive statistics"),
        ("08_continuous_descriptives.csv", "08 Continuous",
         "Continuous variable descriptive statistics"),
        ("09_pearson_correlation.csv", "09 Pearson",
         "Pearson product-moment correlations (continuous and discrete)"),
        ("10_spearman_correlation.csv", "10 Spearman",
         "Spearman rank correlations (ordinal x ordinal)"),
    ]

    table_num = 2
    for csv_fname, sheet_name, title in static_sheets:
        csv_path = os.path.join(output_dir, csv_fname)
        if not os.path.isfile(csv_path):
            continue
        df_sheet = pd.read_csv(csv_path)
        if df_sheet.empty:
            continue
        ws = wb.create_sheet(title=sheet_name)
        _write_apa_table(ws, table_num, title, df_sheet,
                         note="Values rounded to 4 decimal places where applicable.")
        table_num += 1

    # Dynamic sheets: grouped, crosstab, text
    dynamic_patterns = [
        ("11_grouped_by_*.csv", "Grouped descriptive statistics: "),
        ("12_crosstab_*.csv", "Cross-tabulation: "),
        ("13_text_*.csv", "Text frequency analysis: "),
    ]
    used_sheet_names: set = {ws.title for ws in wb.worksheets}

    def _unique_sheet_name(candidate: str) -> str:
        """Truncate to 31 chars and deduplicate with a counter suffix."""
        name = candidate[:31]
        if name not in used_sheet_names:
            used_sheet_names.add(name)
            return name
        counter = 2
        while True:
            suffix = f"_{counter}"
            trimmed = candidate[:31 - len(suffix)] + suffix
            if trimmed not in used_sheet_names:
                used_sheet_names.add(trimmed)
                return trimmed
            counter += 1

    for pattern, title_prefix in dynamic_patterns:
        for match_path in sorted(glob.glob(os.path.join(output_dir, pattern))):
            df_match = pd.read_csv(match_path)
            if df_match.empty:
                continue
            base = os.path.splitext(os.path.basename(match_path))[0]
            sheet_name = _unique_sheet_name(base)
            suffix = base.split("_", 2)[-1] if base.count("_") >= 2 else base
            title = title_prefix + suffix.replace("_", " ")
            ws = wb.create_sheet(title=sheet_name)
            _write_apa_table(ws, table_num, title, df_match,
                             note="Values rounded to 4 decimal places where applicable.")
            table_num += 1

    # Temporal
    temporal_path = os.path.join(output_dir, "14_temporal_trends.csv")
    if os.path.isfile(temporal_path):
        df_temporal = pd.read_csv(temporal_path)
        if not df_temporal.empty:
            ws = wb.create_sheet(title="14 Temporal")
            _write_apa_table(ws, table_num, "Temporal trends by calendar month",
                             df_temporal, note="Period aggregated by calendar month.")
            table_num += 1

    out_path = os.path.join(output_dir, "EXPLORATORY_REPORT.xlsx")
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def save_csv(df: pd.DataFrame, path: str) -> None:
    if df is not None and not df.empty:
        df.to_csv(path, index=False)
        print(f"  Saved: {path}")
    else:
        print(f"  Skipped (no data): {path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Universal EDA runner -- q-exploratory-analysis"
    )
    parser.add_argument("data", help="Path to input file (.xlsx or .csv)")
    parser.add_argument("--output", default="TABLE/", help="Output directory (default: TABLE/)")
    parser.add_argument("--group", nargs="*", default=[], help="Nominal columns for grouping")
    parser.add_argument(
        "--col_types", nargs="*", default=[],
        help="Confirmed column types: col1=type1 col2=type2. "
             "Valid types: id, binary, nominal, ordinal, discrete, continuous, temporal, text.",
    )
    parser.add_argument("--top_n", type=int, default=10, help="Top-N for frequency tables")
    parser.add_argument(
        "--interactive", action="store_true",
        help="Prompt for ambiguous integers (standalone CLI only).",
    )
    parser.add_argument(
        "--no_excel", action="store_true",
        help="Skip Phase 6 (Excel report). CSVs only.",
    )
    args = parser.parse_args()

    col_types_override = {}
    for item in args.col_types:
        col, _, ctype = item.partition("=")
        col_types_override[col] = ctype

    os.makedirs(args.output, exist_ok=True)

    # --- Load data ---
    print(f"\nLoading: {args.data}")
    if args.data.endswith(".xlsx"):
        df = pd.read_excel(args.data, engine="openpyxl")
    elif args.data.endswith(".csv"):
        df = pd.read_csv(args.data)
    else:
        print("ERROR: Input must be .xlsx or .csv")
        sys.exit(1)
    print(f"Loaded {len(df):,} rows x {len(df.columns)} columns.\n")

    # --- Phase 0: Classification ---
    col_types = classify_columns(df, col_types_override=col_types_override)
    col_types = resolve_ambiguous_columns(col_types, df,
                                          no_interactive=not args.interactive)
    print_schema_summary(df, col_types)

    output_files = []

    # --- Phase 1: Dataset Profile ---
    print("=== Phase 1: Dataset Profile ===")
    profile = dataset_profile(df, col_types)
    path = os.path.join(args.output, "01_dataset_profile.csv")
    save_csv(profile, path)
    output_files.append(path)

    # --- Phase 2: Data Quality ---
    print("\n=== Phase 2: Data Quality ===")
    dq = check_data_quality(df, col_types)
    path = os.path.join(args.output, "02_data_quality.csv")
    save_csv(dq, path)
    output_files.append(path)

    # --- Phase 3: Univariate ---
    print("\n=== Phase 3: Univariate Analysis ===")

    nf = nominal_frequencies(df, col_types, top_n=args.top_n)
    path = os.path.join(args.output, "03_nominal_frequencies.csv")
    save_csv(nf, path)
    output_files.append(path)

    bs = binary_summary(df, col_types)
    path = os.path.join(args.output, "04_binary_summary.csv")
    save_csv(bs, path)
    output_files.append(path)

    of = ordinal_frequency(df, col_types)
    path = os.path.join(args.output, "05_ordinal_distribution.csv")
    save_csv(of, path)
    output_files.append(path)

    ordinal_cols = [c for c, t in col_types.items() if t == "ordinal"]
    if ordinal_cols:
        ord_desc = pd.DataFrame([quantitative_summary(df[c], c, "ordinal") for c in ordinal_cols])
        path = os.path.join(args.output, "06_ordinal_descriptives.csv")
        save_csv(ord_desc, path)
        output_files.append(path)

    discrete_cols = [c for c, t in col_types.items() if t == "discrete"]
    if discrete_cols:
        disc_desc = pd.DataFrame([quantitative_summary(df[c], c, "discrete") for c in discrete_cols])
        path = os.path.join(args.output, "07_discrete_descriptives.csv")
        save_csv(disc_desc, path)
        output_files.append(path)

    cont_cols = [c for c, t in col_types.items() if t == "continuous"]
    if cont_cols:
        cont_desc = pd.DataFrame([quantitative_summary(df[c], c, "continuous") for c in cont_cols])
        path = os.path.join(args.output, "08_continuous_descriptives.csv")
        save_csv(cont_desc, path)
        output_files.append(path)

    # --- Phase 4: Bivariate ---
    print("\n=== Phase 4: Bivariate & Multivariate Analysis ===")

    pearson = pearson_correlation_matrix(df, col_types)
    path = os.path.join(args.output, "09_pearson_correlation.csv")
    save_csv(pearson, path)
    output_files.append(path)

    spearman = spearman_correlation_matrix(df, col_types)
    path = os.path.join(args.output, "10_spearman_correlation.csv")
    save_csv(spearman, path)
    output_files.append(path)

    nominal_cols = [c for c, t in col_types.items() if t == "nominal"]
    measure_types = ("continuous", "discrete", "ordinal")
    group_cols = args.group if args.group else nominal_cols[:2]

    for group_col in group_cols:
        if group_col not in df.columns:
            print(f"  WARNING: group column '{group_col}' not found -- skipping.")
            continue
        measure_target_cols = [c for c, t in col_types.items()
                                if t in measure_types and c != group_col]
        all_grouped = []
        for mc in measure_target_cols:
            mt = col_types[mc]
            grp = grouped_stats_by_nominal(df, mc, group_col, mt)
            all_grouped.append(grp)
        non_empty = [g for g in all_grouped if not g.empty]
        if non_empty:
            combined = pd.concat(non_empty, ignore_index=True)
            safe_name = re.sub(r"[^\w]", "_", group_col)
            path = os.path.join(args.output, f"11_grouped_by_{safe_name}.csv")
            save_csv(combined, path)
            output_files.append(path)

    for col1, col2 in combinations(nominal_cols[:4], 2):
        ct = cross_tabulate(df, col1, col2)
        safe1 = re.sub(r"[^\w]", "_", col1)
        safe2 = re.sub(r"[^\w]", "_", col2)
        path = os.path.join(args.output, f"12_crosstab_{safe1}_x_{safe2}.csv")
        save_csv(ct, path)
        output_files.append(path)

    # --- Phase 5: Specialized ---
    print("\n=== Phase 5: Specialized Analysis ===")

    text_cols = [c for c, t in col_types.items() if t == "text"]
    for tc in text_cols:
        if tc not in df.columns:
            continue
        ta = text_analysis(df[tc], tc, top_n=args.top_n)
        safe_name = re.sub(r"[^\w]", "_", tc)
        path = os.path.join(args.output, f"13_text_{safe_name}.csv")
        save_csv(ta, path)
        output_files.append(path)

    tt = temporal_trends(df, col_types)
    if not tt.empty:
        path = os.path.join(args.output, "14_temporal_trends.csv")
        save_csv(tt, path)
        output_files.append(path)

    # --- Phase 6: Excel Report ---
    if not args.no_excel:
        print("\n=== Phase 6: Excel Report ===")
        try:
            excel_path = generate_excel_report(args.output)
            print(f"  Saved: {excel_path}")
        except ImportError:
            print(
                "  WARNING: openpyxl not installed -- skipping Excel report. "
                "Run: pip install 'openpyxl>=3.1'"
            )

    total = len(output_files)
    suffix = "" if args.no_excel else " + EXPLORATORY_REPORT.xlsx"
    print(f"\nDone. {total} CSV file(s){suffix} written to: {args.output}")
    print(f"Run Post-Script step (see SKILL.md) to generate EXPLORATORY_SUMMARY.md.\n")


if __name__ == "__main__":
    main()
